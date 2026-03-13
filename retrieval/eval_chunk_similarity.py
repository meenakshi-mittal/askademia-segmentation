#!/usr/bin/env python3
"""
Evaluate TA response similarity to retrieved chunks inside askademia segmentation.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import os
import random
import sqlite3
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

import numpy as np
from dotenv import load_dotenv
from openai import AzureOpenAI, OpenAI


DEFAULT_EMBED_MODEL = "text-embedding-3-small"


def _load_env(env_path: Path) -> None:
    if env_path.exists():
        load_dotenv(env_path)


def _infer_deployment_from_env() -> Optional[str]:
    ep = (os.getenv("EMBEDDING_ENDPOINT") or "").strip()
    marker = "/deployments/"
    if not ep or marker not in ep:
        return None
    try:
        return ep.split(marker, 1)[1].split("/", 1)[0].strip() or None
    except Exception:
        return None


def _build_embedding_client(
    *,
    provider: str,
    azure_endpoint: Optional[str],
    api_version: Optional[str],
    embedding_endpoint: Optional[str],
) -> OpenAI:
    env_azure_endpoint = (
        os.getenv("EMBEDDING_ENDPOINT")
        or os.getenv("AZURE_OPENAI_ENDPOINT")
        or os.getenv("azure_endpoint")
    )
    env_api_version = os.getenv("AZURE_OPENAI_API_VERSION") or os.getenv("api_version")
    azure_endpoint = embedding_endpoint or azure_endpoint or env_azure_endpoint
    api_version = api_version or env_api_version

    openai_key = os.getenv("OPENAI_API_KEY")
    azure_key = os.getenv("AZURE_OPENAI_API_KEY") or openai_key

    if provider.lower() == "openai":
        if not openai_key:
            raise RuntimeError("OPENAI_API_KEY is required for provider=openai")
        return OpenAI(api_key=openai_key)
    if provider.lower() == "azure":
        if not azure_endpoint or not api_version:
            raise RuntimeError("For provider=azure set endpoint + api_version")
        if not azure_key:
            raise RuntimeError("AZURE_OPENAI_API_KEY or OPENAI_API_KEY required")
        return AzureOpenAI(
            azure_endpoint=azure_endpoint.rstrip("/"),
            api_key=azure_key,
            api_version=api_version,
        )
    api_key = openai_key or azure_key
    if azure_endpoint:
        if not api_version:
            raise RuntimeError("For Azure OpenAI set api_version")
        return AzureOpenAI(
            azure_endpoint=azure_endpoint.rstrip("/"),
            api_key=api_key,
            api_version=api_version,
        )
    if not api_key:
        raise RuntimeError("Missing API key")
    return OpenAI(api_key=api_key)


def _batched(items: List[str], batch_size: int) -> List[List[str]]:
    return [items[i : i + batch_size] for i in range(0, len(items), batch_size)]


def _truncate(text: str, *, max_chars: int) -> str:
    t = (text or "").strip()
    return t if max_chars <= 0 or len(t) <= max_chars else t[:max_chars]


class EmbeddingCache:
    def __init__(self, path: Path) -> None:
        self.path = path
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.conn = sqlite3.connect(str(path))
        self.conn.execute("CREATE TABLE IF NOT EXISTS embeddings (key TEXT PRIMARY KEY, vector TEXT)")
        self.conn.commit()

    def close(self) -> None:
        self.conn.close()

    @staticmethod
    def make_key(text: str, model: str) -> str:
        return f"{model}:{hashlib.sha256(text.encode('utf-8')).hexdigest()}"

    def get_many(self, keys: List[str]) -> Dict[str, List[float]]:
        out: Dict[str, List[float]] = {}
        cur = self.conn.cursor()
        for key in keys:
            cur.execute("SELECT vector FROM embeddings WHERE key = ?", (key,))
            row = cur.fetchone()
            if not row:
                continue
            try:
                out[key] = json.loads(row[0])
            except Exception:
                pass
        return out

    def put_many(self, items: Dict[str, List[float]]) -> None:
        if not items:
            return
        cur = self.conn.cursor()
        for key, vec in items.items():
            cur.execute("INSERT OR REPLACE INTO embeddings (key, vector) VALUES (?, ?)", (key, json.dumps(vec)))
        self.conn.commit()


def _compute_embeddings_cached(
    client: OpenAI,
    texts: List[str],
    *,
    model: str,
    batch_size: int,
    cache: EmbeddingCache,
) -> np.ndarray:
    if not texts:
        return np.asarray([], dtype=np.float32)
    keys = [EmbeddingCache.make_key(t, model) for t in texts]
    cached = cache.get_many(keys)
    missing = [i for i, k in enumerate(keys) if k not in cached]
    if missing:
        for batch in _batched([texts[i] for i in missing], max(1, int(batch_size))):
            resp = client.embeddings.create(model=model, input=batch)
            new_items = {
                EmbeddingCache.make_key(t, model): list(d.embedding)
                for t, d in zip(batch, resp.data)
            }
            cache.put_many(new_items)
            cached.update(new_items)
    ordered = [cached[k] for k in keys]
    return np.asarray(ordered, dtype=np.float32)


def _safe_str(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _is_missing(v: Any) -> bool:
    s = _safe_str(v)
    return s == "" or s.lower() == "nan"


@dataclass(frozen=True)
class QuestionRow:
    sheet: str
    row_idx_1based: int
    lecture_id: str
    question: str
    ta_response: str


def _load_questions_from_xlsx(path: Path) -> List[QuestionRow]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: List[QuestionRow] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        it = ws.iter_rows(values_only=True)
        try:
            header = next(it)
        except StopIteration:
            continue
        if not header:
            continue
        col_to_idx = {_safe_str(c): i for i, c in enumerate(header) if not _is_missing(c)}

        def get(row: Sequence[Any], col: str) -> Any:
            i = col_to_idx.get(col)
            return row[i] if i is not None and i < len(row) else None

        for i0, row in enumerate(it, start=2):
            q = _safe_str(get(row, "Question"))
            if not q:
                continue
            ta = _safe_str(get(row, "TA Response"))
            lecture = _safe_str(get(row, "Lecture"))
            lecture_id = lecture.replace("Lec", "").replace("lec", "").strip()
            if not lecture_id:
                lecture_id = sheet_name.replace("Lec", "").replace("lec", "").strip()
            out.append(
                QuestionRow(
                    sheet=sheet_name,
                    row_idx_1based=i0,
                    lecture_id=lecture_id,
                    question=q,
                    ta_response=ta,
                )
            )
    return out


def _method_folders(data_dir: Path) -> Dict[str, Path]:
    skip = {"audio_sentence_level", "video_ocr", "audio_video_sentence_level", "hashes"}
    out: Dict[str, Path] = {}
    for p in sorted([d for d in data_dir.iterdir() if d.is_dir()], key=lambda d: d.name):
        if p.name in skip:
            continue
        out[p.name] = p
    return out


def _load_chunk_texts(path: Path, *, max_chars: int) -> List[str]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, list):
        return []
    texts: List[str] = []
    for ch in data:
        if not isinstance(ch, dict):
            continue
        txt = str(ch.get("text", "") or "").strip()
        if txt:
            texts.append(_truncate(txt, max_chars=max_chars))
    return texts


def _load_tree_texts(path: Path, *, max_chars: int) -> List[str]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        return []
    roots = data.get("roots")
    if not isinstance(roots, list):
        return []
    texts: List[str] = []

    def visit(node: Dict[str, Any]) -> None:
        title = str(node.get("title", "") or "").strip()
        summary = str(node.get("summary", "") or "").strip()
        parts = [p for p in [title, summary] if p]
        if parts:
            texts.append(_truncate("\n\n".join(parts), max_chars=max_chars))
        leaf = node.get("leaf_chunk")
        if isinstance(leaf, dict):
            lt = str(leaf.get("text", "") or "").strip()
            if lt:
                texts.append(_truncate(lt, max_chars=max_chars))
        for child in node.get("children") or []:
            if isinstance(child, dict):
                visit(child)

    for r in roots:
        if isinstance(r, dict):
            visit(r)
    return texts


def _load_precomputed(
    precomputed_dir: Path,
    *,
    method: str,
    lecture_id: str,
    expected_model: str,
) -> Optional[Tuple[List[str], np.ndarray]]:
    method_dir = precomputed_dir / method
    emb_path = method_dir / f"lecture{lecture_id}.emb.npy"
    txt_path = method_dir / f"lecture{lecture_id}.texts.json"
    meta_path = method_dir / f"lecture{lecture_id}.meta.json"
    if not (emb_path.exists() and txt_path.exists() and meta_path.exists()):
        return None
    try:
        meta = json.loads(meta_path.read_text(encoding="utf-8"))
        if str(meta.get("embedding_model", "")).strip() != expected_model:
            return None
        texts = json.loads(txt_path.read_text(encoding="utf-8"))
        embs = np.load(emb_path)
    except Exception:
        return None
    if not isinstance(texts, list) or len(texts) != len(embs):
        return None
    return texts, embs


def _ci95(vals: List[float]) -> float:
    if len(vals) <= 1:
        return 0.0
    arr = np.asarray(vals, dtype=np.float64)
    return 1.96 * float(np.std(arr, ddof=1)) / math.sqrt(len(vals))


def main() -> None:
    repo_root = Path(__file__).resolve().parents[2]
    seg_dir = repo_root / "askademia segmentation"
    default_env = repo_root / "config" / "keys.env"
    default_xlsx = seg_dir / "ta_responses" / "fa24_labeled_responses.xlsx"

    parser = argparse.ArgumentParser(description="Evaluate retrieval similarity in askademia segmentation.")
    parser.add_argument("--xlsx", type=str, default=str(default_xlsx))
    parser.add_argument("--data-dir", type=str, default=str(seg_dir / "data"))
    parser.add_argument("--out", type=str, default="")
    parser.add_argument("--env", type=str, default=str(default_env))
    parser.add_argument("--provider", type=str, choices=["auto", "openai", "azure"], default="auto")
    parser.add_argument("--model", type=str, default="")
    parser.add_argument("--batch-size", type=int, default=64)
    parser.add_argument("--max-embed-chars", type=int, default=12000)
    parser.add_argument("--top-k", type=int, default=3)
    parser.add_argument("--write-per-row", action="store_true")
    parser.add_argument("--store-retrievals", action="store_true")
    parser.add_argument("--progress-every", type=int, default=20)
    parser.add_argument("--sample", type=int, default=0)
    parser.add_argument("--seed", type=int, default=13)
    parser.add_argument("--precomputed-dir", type=str, default=str(seg_dir / ".cache" / "precomputed_embeddings"))
    parser.add_argument("--no-use-precomputed", action="store_true")
    parser.add_argument("--lecture", type=str, default="")
    parser.add_argument("--cache-path", type=str, default=str(seg_dir / ".cache" / "embeddings.sqlite"))
    parser.add_argument("--generate-responses", action="store_true", help="Accepted for compatibility; no-op.")
    args = parser.parse_args()

    _load_env(Path(args.env))
    emb_model = (args.model or "").strip() or _infer_deployment_from_env() or DEFAULT_EMBED_MODEL
    emb_client = _build_embedding_client(
        provider=args.provider,
        azure_endpoint=None,
        api_version=None,
        embedding_endpoint=None,
    )
    emb_cache = EmbeddingCache(Path(args.cache_path))

    rows = _load_questions_from_xlsx(Path(args.xlsx))
    if args.lecture.strip():
        wanted = {x.strip() for x in args.lecture.split(",") if x.strip()}
        rows = [r for r in rows if str(r.lecture_id) in wanted]
    if int(args.sample) > 0 and len(rows) > int(args.sample):
        rng = random.Random(int(args.seed))
        rows = rng.sample(rows, int(args.sample))
    if not rows:
        raise RuntimeError("No usable rows found for evaluation")

    methods = _method_folders(Path(args.data_dir))
    if not methods:
        raise RuntimeError(f"No method folders found under {args.data_dir}")

    chunk_cache: Dict[Tuple[str, str], Tuple[List[str], np.ndarray]] = {}

    def get_chunks(method: str, lecture_id: str) -> Optional[Tuple[List[str], np.ndarray]]:
        key = (method, lecture_id)
        if key in chunk_cache:
            return chunk_cache[key]
        if not bool(args.no_use_precomputed):
            pre = _load_precomputed(Path(args.precomputed_dir), method=method, lecture_id=lecture_id, expected_model=emb_model)
            if pre is not None:
                chunk_cache[key] = pre
                return pre
        src = methods[method] / f"lecture{lecture_id}.json"
        if not src.exists():
            return None
        texts = _load_tree_texts(src, max_chars=int(args.max_embed_chars)) if "tree" in method.lower() else _load_chunk_texts(src, max_chars=int(args.max_embed_chars))
        if not texts:
            return None
        embs = _compute_embeddings_cached(
            emb_client,
            texts,
            model=emb_model,
            batch_size=int(args.batch_size),
            cache=emb_cache,
        )
        chunk_cache[key] = (texts, embs)
        return chunk_cache[key]

    sums: Dict[str, Dict[str, float]] = {}
    counts: Dict[str, int] = {}
    samples: Dict[str, Dict[str, List[float]]] = {}
    per_row: List[Dict[str, Any]] = []

    total = len(rows)
    for i, row in enumerate(rows, start=1):
        if not row.question or not row.ta_response:
            continue
        q = _truncate(row.question, max_chars=int(args.max_embed_chars))
        ta = _truncate(row.ta_response, max_chars=int(args.max_embed_chars))
        q_emb = _compute_embeddings_cached(emb_client, [q], model=emb_model, batch_size=1, cache=emb_cache)[0]
        ta_emb = _compute_embeddings_cached(emb_client, [ta], model=emb_model, batch_size=1, cache=emb_cache)[0]

        row_scores: Dict[str, Dict[str, float]] = {}
        row_retrievals: Dict[str, List[str]] = {}
        for method in methods:
            payload = get_chunks(method, row.lecture_id)
            if payload is None:
                continue
            texts, chunk_embs = payload
            if len(chunk_embs) == 0:
                continue
            sims = np.dot(chunk_embs, q_emb) / (np.linalg.norm(chunk_embs, axis=1) * np.linalg.norm(q_emb) + 1e-12)
            k = min(int(args.top_k), len(sims))
            if k <= 0:
                continue
            top_idx = np.argpartition(-sims, k - 1)[:k]
            top_embs = chunk_embs[top_idx]
            ta_sims = np.dot(top_embs, ta_emb) / (np.linalg.norm(top_embs, axis=1) * np.linalg.norm(ta_emb) + 1e-12)
            max_cos = float(np.max(ta_sims))
            avg_cos = float(np.mean(ta_sims))
            row_scores[method] = {"max_cosine": max_cos, "avg_cosine": avg_cos}
            if bool(args.store_retrievals):
                row_retrievals[method] = [texts[int(j)] for j in top_idx]

            sums.setdefault(method, {"max_cosine": 0.0, "avg_cosine": 0.0})
            counts[method] = counts.get(method, 0) + 1
            sums[method]["max_cosine"] += max_cos
            sums[method]["avg_cosine"] += avg_cos
            samples.setdefault(method, {"max_cosine": [], "avg_cosine": []})
            samples[method]["max_cosine"].append(max_cos)
            samples[method]["avg_cosine"].append(avg_cos)

        if row_scores and bool(args.write_per_row or args.store_retrievals):
            out_row: Dict[str, Any] = {
                "sheet": row.sheet,
                "row_idx_1based": row.row_idx_1based,
                "lecture_id": row.lecture_id,
                "question": row.question,
                "ta_response": row.ta_response,
                "scores": row_scores,
            }
            if bool(args.store_retrievals):
                out_row["retrievals"] = row_retrievals
            per_row.append(out_row)

        pe = int(args.progress_every)
        if pe > 0 and (i == 1 or i % pe == 0):
            print(f"progress: {i}/{total}")

    averages: Dict[str, Dict[str, float]] = {}
    for method, agg in sums.items():
        n = counts.get(method, 0)
        if n <= 0:
            continue
        max_vals = samples.get(method, {}).get("max_cosine", [])
        avg_vals = samples.get(method, {}).get("avg_cosine", [])
        averages[method] = {
            "n": n,
            "avg_max_cosine": agg["max_cosine"] / n,
            "avg_avg_cosine": agg["avg_cosine"] / n,
            "ci95_avg_max_cosine": _ci95(max_vals),
            "ci95_avg_avg_cosine": _ci95(avg_vals),
        }

    out_payload: Dict[str, Any] = {
        "xlsx": str(args.xlsx),
        "embedding_model": emb_model,
        "top_k": int(args.top_k),
        "averages": averages,
    }
    if bool(args.write_per_row or args.store_retrievals):
        out_payload["per_row"] = per_row

    if args.out:
        out_path = Path(args.out)
    else:
        stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        out_path = seg_dir / "retrieval" / f"ta_similarity_{stamp}.json"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(out_payload, ensure_ascii=False, indent=2), encoding="utf-8")
    emb_cache.close()
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()

